"""Excel extraction with formula parsing and dependency analysis."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ingest.base import PageResult

CELL_REF_RE = re.compile(r"(?<![A-Za-z])([A-Z]{1,3})(\d{1,7})(?!\d)")
RANGE_REF_RE = re.compile(r"([A-Z]{1,3}\d+):([A-Z]{1,3}\d+)")


def _col_index(col_letter: str) -> int:
    result = 0
    for c in col_letter:
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def _get_headers(ws) -> dict[int, str]:
    """Map column index → header text from row 1."""
    headers: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            headers[col] = str(val).strip()
    return headers


def _format_cell_value(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _format_sheet_as_table(ws, ws_values) -> str:
    """Render sheet data as a readable text table with formulas noted."""
    if ws.max_row is None or ws.max_column is None:
        return ""
    if ws.max_row < 1 or ws.max_column < 1:
        return ""

    lines: list[str] = [f"Sheet: \"{ws.title}\""]

    headers = _get_headers(ws)
    if headers:
        header_line = " | ".join(headers.get(c, f"Col{c}") for c in range(1, ws.max_column + 1))
        lines.append(header_line)
        lines.append("-" * len(header_line))

    max_rows = min(ws.max_row, 200)
    start_row = 2 if headers else 1

    for row in range(start_row, max_rows + 1):
        cells: list[str] = []
        for col in range(1, ws.max_column + 1):
            formula_cell = ws.cell(row=row, column=col)
            value_cell = ws_values.cell(row=row, column=col) if ws_values else formula_cell
            formula_val = formula_cell.value
            computed_val = value_cell.value if ws_values else None

            if isinstance(formula_val, str) and formula_val.startswith("="):
                if computed_val is not None:
                    cells.append(f"{formula_val} → {_format_cell_value(computed_val)}")
                else:
                    cells.append(formula_val)
            else:
                cells.append(_format_cell_value(formula_val))
        if any(c for c in cells):
            lines.append(" | ".join(cells))

    return "\n".join(lines)


def _extract_formulas(ws, headers: dict[int, str]) -> list[dict]:
    """Extract all formula cells with their references."""
    formulas: list[dict] = []
    if ws.max_row is None or ws.max_column is None:
        return formulas

    for row in range(1, min(ws.max_row + 1, 500)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue

            coord = f"{get_column_letter(col)}{row}"
            col_header = headers.get(col, get_column_letter(col))

            refs = CELL_REF_RE.findall(val)
            ref_descriptions: list[str] = []
            for ref_col, ref_row in refs:
                ref_col_idx = _col_index(ref_col)
                ref_header = headers.get(ref_col_idx, ref_col)
                ref_descriptions.append(f"{ref_header} ({ref_col}{ref_row})")

            ranges = RANGE_REF_RE.findall(val)
            for start, end in ranges:
                start_col = re.match(r"[A-Z]+", start).group()
                start_col_idx = _col_index(start_col)
                range_header = headers.get(start_col_idx, start_col)
                ref_descriptions.append(f"{range_header} range ({start}:{end})")

            formulas.append({
                "cell": coord,
                "column_header": col_header,
                "formula": val,
                "references": ref_descriptions,
            })

    return formulas


def _format_formula_summary(ws_title: str, formulas: list[dict]) -> str:
    """Produce a human-readable summary of formula logic."""
    if not formulas:
        return ""

    lines = [f"Formula Logic — Sheet \"{ws_title}\":"]
    for f in formulas:
        refs = ", ".join(f["references"]) if f["references"] else "constants"
        lines.append(
            f"- \"{f['column_header']}\" ({f['cell']}) {f['formula']}  ← depends on: {refs}"
        )

    return "\n".join(lines)


def extract(path: Path) -> PageResult:
    wb_formulas = load_workbook(str(path), data_only=False, read_only=False)
    try:
        wb_values = load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        wb_values = None

    pages: PageResult = []

    for sheet_idx, ws in enumerate(wb_formulas.worksheets, start=1):
        ws_values = wb_values[ws.title] if wb_values and ws.title in wb_values.sheetnames else None

        table_text = _format_sheet_as_table(ws, ws_values)
        if not table_text.strip():
            continue

        headers = _get_headers(ws)
        formulas = _extract_formulas(ws, headers)
        formula_summary = _format_formula_summary(ws.title, formulas)

        parts = [table_text]
        if formula_summary:
            parts.append(formula_summary)

        pages.append((sheet_idx, "\n\n".join(parts)))

    wb_formulas.close()
    if wb_values:
        wb_values.close()

    return pages
